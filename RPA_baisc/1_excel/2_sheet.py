from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() #새로운 시트 기본 이름으로 설정
ws.title = "MY sheet"
ws1 = wb.create_sheet("Yoursheet")
ws2 = wb.create_sheet("you sheet", 2) #2번째 인덱스에 시트를 삽입

new_ws = wb["you sheet"] #dict형태로 sheet에 접근

print(wb.sheetnames) #모든 sheet 이름확인

#sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "copied sheet"


#sheet, mysheet, yoursheet
wb.save("sample.xlsx")
wb.close()